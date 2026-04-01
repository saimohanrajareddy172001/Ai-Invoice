#!/usr/bin/env python3
"""
Export invoice line items for a date range to Excel.
Usage: python3 tools/export_invoice_lines.py 2026-03-23 2026-03-29
Output: .tmp/invoice_lines_<start>_to_<end>.xlsx
"""

import sys
import os
import json
import urllib.request
import urllib.parse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY")


def fetch_invoice_lines(start_date: str, end_date: str) -> list:
    auth_headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }

    # Step 1: Get header IDs for the date range
    params = urllib.parse.urlencode({
        "select": "id,invoice_number,vendor,invoice_date",
        "invoice_date": f"gte.{start_date}",
        "order": "invoice_date.desc",
    })
    url = f"{SUPABASE_URL}/rest/v1/invoice_headers?{params}&invoice_date=lte.{end_date}"
    req = urllib.request.Request(url, headers=auth_headers)
    with urllib.request.urlopen(req) as r:
        hdrs = json.loads(r.read().decode())

    if not hdrs:
        return []

    # Build a lookup map
    hdr_map = {h["id"]: h for h in hdrs}
    hdr_ids = ",".join(str(h["id"]) for h in hdrs)

    # Step 2: Fetch all invoice_lines for those header IDs
    params2 = urllib.parse.urlencode({
        "select": "header_id,item_name,category,unit_qty,case_qty,unit_price,total",
        "header_id": f"in.({hdr_ids})",
        "order": "item_name.asc",
        "limit": "10000",
    })
    url2 = f"{SUPABASE_URL}/rest/v1/invoice_lines?{params2}"
    req2 = urllib.request.Request(url2, headers=auth_headers)
    with urllib.request.urlopen(req2) as r:
        lines = json.loads(r.read().decode())

    # Attach header info to each line
    for line in lines:
        hdr = hdr_map.get(line["header_id"], {})
        line["invoice_date"] = hdr.get("invoice_date", "")
        line["invoice_number"] = hdr.get("invoice_number", "")
        line["vendor"] = hdr.get("vendor", "")

    # Sort by date desc, then item name
    lines.sort(key=lambda x: (x["invoice_date"], x["item_name"]), reverse=False)
    lines.sort(key=lambda x: x["invoice_date"], reverse=True)
    return lines


def export_to_excel(rows: list, start_date: str, end_date: str) -> str:
    os.makedirs(".tmp", exist_ok=True)
    out_path = f".tmp/invoice_lines_{start_date}_to_{end_date}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Lines"

    # Header row styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    columns = ["Date", "Invoice #", "Vendor", "Item", "Category", "Unit Qty", "Case Qty", "Unit Price", "Total"]
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.get("invoice_date", ""))
        ws.cell(row=row_idx, column=2, value=row.get("invoice_number", ""))
        ws.cell(row=row_idx, column=3, value=row.get("vendor", ""))
        ws.cell(row=row_idx, column=4, value=row.get("item_name", ""))
        ws.cell(row=row_idx, column=5, value=row.get("category", ""))
        ws.cell(row=row_idx, column=6, value=row.get("unit_qty", 0))
        ws.cell(row=row_idx, column=7, value=row.get("case_qty", 0))
        ws.cell(row=row_idx, column=8, value=row.get("unit_price", 0))
        ws.cell(row=row_idx, column=9, value=row.get("total", 0))

    # Auto-fit column widths
    for col_idx in range(1, len(columns) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Total row at bottom
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=8, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=f"=SUM(I2:I{total_row - 1})").font = Font(bold=True)

    wb.save(out_path)
    return out_path


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 tools/export_invoice_lines.py <start_date> <end_date>")
        print("Example: python3 tools/export_invoice_lines.py 2026-03-23 2026-03-29")
        sys.exit(1)

    start_date, end_date = sys.argv[1], sys.argv[2]

    # Validate dates
    for d in [start_date, end_date]:
        datetime.strptime(d, "%Y-%m-%d")

    print(f"Fetching invoice lines from {start_date} to {end_date}...")
    rows = fetch_invoice_lines(start_date, end_date)
    print(f"Found {len(rows)} line items")

    if not rows:
        print("No data found for this date range.")
        sys.exit(0)

    out_path = export_to_excel(rows, start_date, end_date)
    print(f"Saved to: {out_path}")


if __name__ == "__main__":
    main()
